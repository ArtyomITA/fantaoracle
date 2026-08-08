# -*- coding: utf-8 -*-
"""Scansione rapida dei file extra: fogli, presenza dati d'asta, stima stagione."""
import glob
import os

from openpyxl import load_workbook

from ge_parser import load_grid, find_anchors, parse_sheet

EXTRA_DIR = r"data\raw\gruppoesperti\extra"

# marker per stima stagione: giocatori presenti in Serie A SOLO (o quasi) in quella stagione
SEASON_MARKERS = {
    "2021-22": ["HANDANOVIC", "VIDAL", "OSPINA", "IBRAHIMOVIC", "TATARUSANU",
                "STRAKOSHA", "MURIEL", "CALLEJON", "RIBERY", "MERTENS", "INSIGNE"],
    "2022-23": ["ONANA", "KIM", "DIA", "CDK", "BROZOVIC", "SKRINIAR",
                "ISMAJLI", "TERRACCIANO", "OSIMHEN", "KVARATSKHELIA"],
    "2023-24": ["RETEGUI", "THURAM", "GIROUD", "FRATTESI", "COLPANI",
                "GUDMUNDSSON", "ZIRKZEE", "SOULE", "OSIMHEN", "KVARATSKHELIA"],
    "2024-25": ["DOVBYK", "RETEGUI", "LUKAKU", "THURAM", "MORATA",
                "KVARATSKHELIA", "PULISIC", "COLPANI", "ZAPATA", "CASTRO"],
    "2025-26": ["MODRIC", "DAVID", "OPENDA", "DZEKO", "IMMOBILE",
                "DE BRUYNE", "VLAHOVIC", "PULISIC", "LAUTARO", "LEONI"],
}

for path in sorted(glob.glob(os.path.join(EXTRA_DIR, "*.xlsx"))):
    name = os.path.basename(path)
    size = os.path.getsize(path)
    print("=" * 78)
    print(f"{name}  ({size} byte)")
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print("  ERRORE apertura:", e)
        continue
    print("  fogli:", wb.sheetnames)
    wb.close()

    grid = load_grid(path, "Aste Concluse")
    if grid is None:
        # cerca COMPONENTI in tutti i fogli
        wb = load_workbook(path, data_only=True, read_only=True)
        hits = {}
        for sn in wb.sheetnames:
            g = [[c.value for c in row] for row in wb[sn].iter_rows(max_row=300)]
            n = sum(1 for row in g for v in row
                    if isinstance(v, str) and v.strip().upper() == "COMPONENTI")
            if n:
                hits[sn] = n
        wb.close()
        print("  nessun foglio 'Aste Concluse'; COMPONENTI trovato in:", hits or "nessuno")
        continue

    anchors = find_anchors(grid)
    rows, stats = parse_sheet(grid, name)
    print(f"  'Aste Concluse': ancore={len(anchors)} aste_popolate={stats['populated_auctions']} righe={len(rows)}")

    names = {str(r["player_raw"]).strip().upper() for r in rows}
    scores = {}
    for season, markers in SEASON_MARKERS.items():
        scores[season] = sum(1 for m in markers if m in names)
    print("  marker stagione:", scores)
    sample = sorted(names)[:0]  # placeholder
    # mostra qualche nome distintivo
    for probe in ["DOVBYK", "KVARATSKHELIA", "RETEGUI", "HANDANOVIC", "VIDAL",
                  "IBRAHIMOVIC", "ONANA", "GIROUD", "ZIRKZEE", "MODRIC", "DE BRUYNE",
                  "LUKAKU", "OSIMHEN", "IMMOBILE", "THURAM"]:
        if probe in names:
            print("    presente:", probe)
