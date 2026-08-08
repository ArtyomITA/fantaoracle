# -*- coding: utf-8 -*-
"""Scansione ancore 'COMPONENTI' / 'ASTA n' nel foglio 'Aste Concluse'."""
from openpyxl import load_workbook

FILES = [
    r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2024-25.xlsx",
    r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2021-22circa.xlsx",
]

for path in FILES:
    print("=" * 80)
    print(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Aste Concluse"]
    anchors = []       # (row, col) of COMPONENTI
    asta_labels = []   # (row, col, label)
    role_headers = {}  # role -> count, cols set
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            s = v.strip().upper()
            if s == "COMPONENTI":
                anchors.append((c.row, c.column))
            elif s.startswith("ASTA "):
                asta_labels.append((c.row, c.column, v.strip()))
            elif s in ("PORTIERI", "DIFENSORI", "CENTROCAMPISTI", "ATTACCANTI"):
                role_headers.setdefault(s, []).append((c.row, c.column))
    print(f"COMPONENTI anchors: {len(anchors)}")
    print("  colonne distinte:", sorted({a[1] for a in anchors}))
    print("  prime 5:", anchors[:5])
    print("  ultime 5:", anchors[-5:])
    rows = [a[0] for a in anchors]
    gaps = [rows[i+1] - rows[i] for i in range(len(rows)-1)]
    import collections
    print("  gap fra ancore consecutive:", collections.Counter(gaps).most_common(10))
    print(f"ASTA labels: {len(asta_labels)}; prime 3: {asta_labels[:3]}; ultime 3: {asta_labels[-3:]}")
    for r, lst in sorted(role_headers.items()):
        print(f"  header {r}: n={len(lst)} colonne={sorted({x[1] for x in lst})}")
    wb.close()
