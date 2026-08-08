# -*- coding: utf-8 -*-
"""Ispezione fogli aggregati (PORTIERI ecc.) per capire quante aste contengono (valori cached)."""
from openpyxl import load_workbook

FILES = {
    "2024-25": r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2024-25.xlsx",
    "2021-22": r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2021-22circa.xlsx",
}

for label, path in FILES.items():
    print("=" * 80, label)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["PORTIERI"]
    rows = [[c.value for c in row] for row in ws.iter_rows(max_row=30)]
    # stampa le prime 12 righe, prime 30 colonne, solo celle non vuote
    for ri, r in enumerate(rows[:12], start=1):
        vals = [(ci + 1, v) for ci, v in enumerate(r[:36]) if v is not None]
        if vals:
            print(f"R{ri}: {vals}")
    # quante colonne hanno dati nella riga di intestazione asta?
    wb.close()
