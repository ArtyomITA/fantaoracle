# -*- coding: utf-8 -*-
"""Parser generico per i fogli 'Aste Concluse' del Progetto Prezzi Asta (GruppoEsperti).

Layout accertato (verificato a mano su blocchi 1, 2, 11, 140 di entrambi i file):
- blocchi-asta impilati verticalmente con passo 110 righe, tutti a colonna 1;
- ancora del blocco = cella con etichetta 'COMPONENTI' (riga r, colonna k);
  'ASTA n' a (r-1,k); 'PERIODO' (r+1,k); 'MODIFICATORE' (r+2,k); 'CREDITI TOT' (r+3,k);
  i valori stanno 2 colonne a destra dell'etichetta (colonna k+2);
- riga intestazioni ruolo a r+6: 'PORTIERI' a k, 'DIFENSORI' a k+6,
  'CENTROCAMPISTI' a k+12, 'ATTACCANTI' a k+18; per ogni sezione-ruolo con
  etichetta in colonna c0: numero=c0, nome_raw=c0+1, crediti=c0+2, %=c0+3,
  nome_normalizzato=c0+4;
- righe giocatori da r+7 fino alla fine del blocco (r+108 col passo standard).

Il parser non assume il passo: scopre le ancore programmaticamente e delimita
ogni blocco con l'ancora successiva nella stessa colonna (robusto anche a
eventuali blocchi affiancati in orizzontale).
"""
from openpyxl import load_workbook

ROLE_MAP = {
    "PORTIERI": "P",
    "DIFENSORI": "D",
    "CENTROCAMPISTI": "C",
    "ATTACCANTI": "A",
}
BLOCK_WIDTH = 24  # larghezza massima blocco (colonne)


def _num(v):
    """Converte in int se intero, altrimenti float; None se non numerico."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f.is_integer() else f
    s = str(v).strip().replace(",", ".")
    if s == "":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def load_grid(path, sheet_name="Aste Concluse"):
    """Carica il foglio come lista di liste (valori cached, data_only)."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet_name]
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    return grid


def _cell(grid, r, c):
    """r, c 1-based."""
    if r < 1 or r > len(grid):
        return None
    row = grid[r - 1]
    if c < 1 or c > len(row):
        return None
    return row[c - 1]


def find_anchors(grid):
    """Trova tutte le celle con etichetta 'COMPONENTI' → lista (riga, colonna) 1-based."""
    anchors = []
    for ri, row in enumerate(grid, start=1):
        for ci, v in enumerate(row, start=1):
            if isinstance(v, str) and v.strip().upper() == "COMPONENTI":
                anchors.append((ri, ci))
    return anchors


def parse_sheet(grid, source_file):
    """Parsa la griglia in righe tidy. Ritorna (rows, stats)."""
    anchors = find_anchors(grid)
    # ancora successiva nella stessa colonna → delimita il blocco in basso
    by_col = {}
    for r, c in anchors:
        by_col.setdefault(c, []).append(r)
    for c in by_col:
        by_col[c].sort()

    rows_out = []
    stats = {
        "anchors": len(anchors),
        "populated_auctions": 0,
        "empty_blocks": 0,
        "rows_no_price": 0,
        "auction_meta": [],  # (auction_id, componenti, crediti_tot, modificatore, periodo, n_righe)
    }
    auction_id = 0
    for r, c in sorted(anchors):
        componenti = _num(_cell(grid, r, c + 2))
        periodo = _num(_cell(grid, r + 1, c + 2))
        modificatore = _cell(grid, r + 2, c + 2)
        crediti_tot = _num(_cell(grid, r + 3, c + 2))
        modificatore = "" if modificatore is None else str(modificatore).strip()

        # riga intestazioni ruolo: cerca tra r+4 e r+10 una riga con 'PORTIERI' nel blocco
        header_row = None
        role_cols = {}
        for hr in range(r + 4, r + 11):
            found = {}
            for cc in range(c, c + BLOCK_WIDTH):
                v = _cell(grid, hr, cc)
                if isinstance(v, str) and v.strip().upper() in ROLE_MAP:
                    found[ROLE_MAP[v.strip().upper()]] = cc
            if "P" in found:
                header_row = hr
                role_cols = found
                break
        if header_row is None:
            stats["empty_blocks"] += 1
            continue

        # fine blocco: ancora successiva nella stessa colonna - 2 (riga prima di 'ASTA n')
        later = [x for x in by_col[c] if x > r]
        end_row = (later[0] - 2) if later else min(r + 108, len(grid))

        players = []
        for pr in range(header_row + 1, end_row + 1):
            for ruolo, c0 in role_cols.items():
                name = _cell(grid, pr, c0 + 1)
                if name is None:
                    continue
                name_s = str(name).strip()
                if name_s == "" or name_s.upper() in ROLE_MAP:
                    continue
                prezzo = _num(_cell(grid, pr, c0 + 2))
                if prezzo is None:
                    stats["rows_no_price"] += 1
                    continue
                players.append((ruolo, name_s, prezzo))

        if not players and componenti is None:
            stats["empty_blocks"] += 1
            continue

        auction_id += 1
        stats["populated_auctions"] += 1
        stats["auction_meta"].append(
            (auction_id, componenti, crediti_tot, modificatore, periodo, len(players))
        )
        for ruolo, name_s, prezzo in players:
            pct = (
                round(prezzo / crediti_tot, 6)
                if (crediti_tot not in (None, 0))
                else None
            )
            rows_out.append(
                {
                    "source_file": source_file,
                    "auction_id": auction_id,
                    "componenti": componenti,
                    "crediti_tot": crediti_tot,
                    "modificatore": modificatore,
                    "periodo": periodo,
                    "ruolo": ruolo,
                    "player_raw": name_s,
                    "prezzo": prezzo,
                    "pct_budget": pct,
                }
            )
    return rows_out, stats


def parse_file(path, source_file, sheet_name="Aste Concluse"):
    grid = load_grid(path, sheet_name)
    if grid is None:
        return None, {"error": f"sheet {sheet_name!r} non presente"}
    return parse_sheet(grid, source_file)
