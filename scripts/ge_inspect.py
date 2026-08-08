# -*- coding: utf-8 -*-
"""Ispezione layout dei fogli 'Aste Concluse' nei file GruppoEsperti."""
import sys
from openpyxl import load_workbook

FILES = [
    r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2024-25.xlsx",
    r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2021-22circa.xlsx",
]

for path in FILES:
    print("=" * 80)
    print(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        try:
            dims = ws.calculate_dimension()
        except ValueError:
            dims = ws.calculate_dimension(force=True)
        print(f"  {name!r}: dims={dims} max_row={ws.max_row} max_col={ws.max_column}")
    wb.close()

# Dettaglio del foglio Aste Concluse del primo file: prime 40 righe x 30 colonne
print("\n\nDETTAGLIO primo file, 'Aste Concluse', righe 1-40, colonne 1-30")
wb = load_workbook(FILES[0], data_only=True, read_only=True)
ws = wb["Aste Concluse"]
for row in ws.iter_rows(min_row=1, max_row=40, max_col=30):
    for c in row:
        if c.value is not None:
            v = str(c.value)
            if len(v) > 25:
                v = v[:25] + "..."
            print(f"  R{c.row}C{c.column}: {v!r}")
    print("  ---")
wb.close()
