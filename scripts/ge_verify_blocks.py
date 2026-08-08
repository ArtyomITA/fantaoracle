# -*- coding: utf-8 -*-
"""Verifica manuale di alcuni blocchi-asta prima di generalizzare."""
from openpyxl import load_workbook

FILES = {
    "2024-25": r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2024-25.xlsx",
    "2021-22": r"E:\claudecode pesante\fonti_prezzi\gruppoesperti_prezzi_aste_reali_2021-22circa.xlsx",
}

def dump_block(ws_rows, asta_idx, label):
    """asta_idx 1-based. Block: rows (asta_idx-1)*110+1 .. asta_idx*110."""
    base = (asta_idx - 1) * 110  # row of 'ASTA n' = base+1
    print(f"--- {label} ASTA {asta_idx} (righe {base+1}-{base+110}) ---")
    hdr = {}
    for off in range(1, 6):
        r = ws_rows[base + off - 1]
        lab = r[0] if len(r) > 0 else None
        val = r[2] if len(r) > 2 else None
        hdr[str(lab)] = val
        print(f"  R{base+off}: label={lab!r} val={val!r}")
    # role header row = base+8, players from base+9
    rrow = ws_rows[base + 8 - 1]
    print(f"  role header row R{base+8}: {[rrow[i] if i < len(rrow) else None for i in (0,2,3,6,8,9,12,14,15,18,20,21)]}")
    # count players per role and show first 2 + last populated
    for role, c0 in (("P", 0), ("D", 6), ("C", 12), ("A", 18)):
        players = []
        for off in range(9, 111):
            if base + off - 1 >= len(ws_rows):
                break
            r = ws_rows[base + off - 1]
            name = r[c0 + 1] if len(r) > c0 + 1 else None
            price = r[c0 + 2] if len(r) > c0 + 2 else None
            if name is not None and str(name).strip() != "":
                players.append((base + off, name, price))
        show = players[:2] + (players[-1:] if len(players) > 2 else [])
        print(f"  {role}: n={len(players)} esempi={show}")

for label, path in FILES.items():
    print("=" * 80)
    print(label)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Aste Concluse"]
    rows = [[c.value for c in row] for row in ws.iter_rows(max_row=22000, max_col=24)]
    wb.close()

    # quali blocchi hanno COMPONENTI valorizzato?
    populated = []
    for i in range(1, 201):
        base = (i - 1) * 110
        comp = rows[base + 1][2] if len(rows[base + 1]) > 2 else None
        # conta giocatori nel blocco
        nplayers = 0
        for off in range(9, 111):
            if base + off - 1 >= len(rows):
                break
            r = rows[base + off - 1]
            for c0 in (0, 6, 12, 18):
                nm = r[c0 + 1] if len(r) > c0 + 1 else None
                if nm is not None and str(nm).strip() != "":
                    nplayers += 1
        if comp is not None or nplayers > 0:
            populated.append((i, comp, nplayers))
    print(f"blocchi con dati: {len(populated)}; ultimi 3: {populated[-3:]}")
    empties_with_players = [p for p in populated if p[1] is None]
    print(f"blocchi con giocatori ma COMPONENTI vuoto: {empties_with_players[:10]}")

    dump_block(rows, 2, label)
    dump_block(rows, populated[-1][0], label)
    if len(populated) < 200:
        dump_block(rows, populated[-1][0] + 1, label + " (dovrebbe essere vuoto)")
