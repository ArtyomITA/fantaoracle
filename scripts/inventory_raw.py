# -*- coding: utf-8 -*-
"""Censimento file grezzi: per ogni CSV/XLSX -> righe, colonne, 3 righe di esempio.
Scrive la sezione inventario in un file intermedio (inventory_section.md) che
inventory_build_md.py assembla in INVENTORY.md insieme a overlap e buchi."""
import os
import sys
import io
import csv
import json

import pandas as pd

ROOTS = [
    r"data\raw",
    r"E:\claudecode pesante\fonti_prezzi",
]
OUT = r"data\raw\_inventory_section.md"

MAX_CELL = 40          # tronca celle lunghe negli esempi
MAX_ROW_CHARS = 220    # tronca la riga di esempio


def fmt_row(values):
    cells = []
    for v in values:
        s = "" if v is None else str(v)
        s = s.replace("\n", " ").replace("\r", " ")
        if len(s) > MAX_CELL:
            s = s[:MAX_CELL - 1] + "…"
        cells.append(s)
    line = " | ".join(cells)
    if len(line) > MAX_ROW_CHARS:
        line = line[:MAX_ROW_CHARS - 1] + "…"
    return line


def sniff_sep(path):
    with open(path, "rb") as f:
        head = f.read(8192)
    txt = head.decode("utf-8", errors="replace")
    first = txt.splitlines()[0] if txt.splitlines() else ""
    if first.count(";") > first.count(","):
        return ";"
    return ","


def inspect_csv(path):
    sep = sniff_sep(path)
    try:
        df = pd.read_csv(path, sep=sep, encoding="utf-8-sig", dtype=str,
                         keep_default_na=False, low_memory=False)
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}"}
    return {
        "rows": len(df),
        "sep": sep,
        "cols": list(df.columns),
        "sample": [fmt_row(df.iloc[i].tolist()) for i in range(min(3, len(df)))],
    }


def inspect_xlsx(path):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}"}
    sheets = []
    for ws in wb.worksheets:
        n_rows = 0
        n_cols = 0
        sample = []
        # scandisce tutte le righe (read_only: max_row/max_column spesso None
        # se il file non dichiara le dimension); campiona le prime 3 non vuote
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            n_rows = i + 1
            n_cols = max(n_cols, len(row))
            if len(sample) < 3 and any(v is not None and str(v).strip() != "" for v in row):
                sample.append(f"r{i+1}: " + fmt_row(row[:30]))
        sheets.append({"name": ws.title, "rows": n_rows, "cols": n_cols, "sample": sample})
    wb.close()
    return {"sheets": sheets}


def main():
    out = io.StringIO()
    tot_csv = tot_xlsx = tot_rows = 0
    other_files = []

    for root in ROOTS:
        out.write(f"\n## Radice: `{root}`\n")
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames.sort()
            filenames.sort()
            rel_dir = os.path.relpath(dirpath, root)
            data_files = [f for f in filenames
                          if f.lower().endswith((".csv", ".xlsx", ".xls"))]
            # gli .xls legacy di xls_originali sono le sorgenti 1:1 dei CSV
            # fantasoccer gia' censiti: li riassumiamo in blocco
            if os.path.basename(dirpath).lower() == "xls_originali":
                xls = [f for f in data_files if f.lower().endswith(".xls")]
                tot_size = sum(os.path.getsize(os.path.join(dirpath, f)) for f in xls)
                out.write(f"\n### Cartella `{rel_dir}`\n\n")
                out.write(f"- **{len(xls)} file .xls legacy** ({tot_size/1024/1024:.1f} MB totali), "
                          f"da `{xls[0]}` a `{xls[-1]}`: sorgenti originali fanta.soccer, "
                          f"contenuto identico ai CSV omonimi in `quotazioni/` (gia' censiti sopra). "
                          f"Formato .xls non leggibile da openpyxl; conservati come backup.\n")
                continue
            for f in filenames:
                if not f.lower().endswith((".csv", ".xlsx", ".xls")):
                    p = os.path.join(dirpath, f)
                    other_files.append((os.path.relpath(p, root), os.path.getsize(p), root))
            if not data_files:
                continue
            out.write(f"\n### Cartella `{rel_dir}`\n\n")
            for f in data_files:
                p = os.path.join(dirpath, f)
                rel = os.path.relpath(p, root)
                size_kb = os.path.getsize(p) / 1024
                if f.lower().endswith(".csv"):
                    info = inspect_csv(p)
                    tot_csv += 1
                    if "error" in info:
                        out.write(f"- **{rel}** ({size_kb:.0f} KB) — ERRORE: {info['error']}\n")
                        continue
                    tot_rows += info["rows"]
                    sep_note = "" if info["sep"] == "," else f" — **separatore `{info['sep']}`**"
                    out.write(f"- **{rel}** ({size_kb:.0f} KB){sep_note}\n")
                    out.write(f"  - righe: **{info['rows']:,}** — colonne ({len(info['cols'])}): "
                              f"`{', '.join(info['cols'])}`\n")
                    for s in info["sample"]:
                        out.write(f"  - es: `{s}`\n")
                else:
                    info = inspect_xlsx(p)
                    tot_xlsx += 1
                    if "error" in info:
                        out.write(f"- **{rel}** ({size_kb:.0f} KB) — ERRORE: {info['error']}\n")
                        continue
                    out.write(f"- **{rel}** ({size_kb:.0f} KB) — xlsx, {len(info['sheets'])} fogli\n")
                    for sh in info["sheets"]:
                        out.write(f"  - foglio \"{sh['name']}\": {sh['rows']:,} righe x {sh['cols']} col\n")
                        for s in sh["sample"]:
                            out.write(f"    - es {s}\n")
    out.write("\n## File non tabellari presenti (supporto)\n\n")
    for rel, size, root in other_files:
        out.write(f"- `{rel}` ({size/1024:.0f} KB) [{os.path.basename(root)}]\n")
    out.write(f"\n**Totali: {tot_csv} CSV ({tot_rows:,} righe dati complessive), {tot_xlsx} XLSX.**\n")

    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write(out.getvalue())
    print(f"OK: {tot_csv} csv, {tot_xlsx} xlsx, {tot_rows:,} righe -> {OUT}")


if __name__ == "__main__":
    main()
